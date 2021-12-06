# This is a sample Python script.
import bs4
import pyautogui
from openpyxl import Workbook
from selenium import webdriver
import tkinter as tk

global username
global password
global browser
global tk_window
file_location = "Esetcustomers.xlsx"
chillisoft_prefix = "https://secure.chillisoft.net"
login_url = "https://secure.chillisoft.net/login/"
postVerifiedUrl = "https://secure.chillisoft.net/ssl/extranet/index.cfm/licence-filter/1/due-to-expire-in-next-6-weeks/"


def run():
    sheets_and_order_setup()


def gui_app(message):  # ----Gets the User's Details----
    global username
    global password
    global browser
    global tk_window
    tk_window = tk.Tk()
    tk_window.geometry('400x150')
    tk_window.title('Chillisoft Login')
    # username label and text entry box
    tk.Label(tk_window, text=message)
    tk.Label(tk_window, text="User Name").grid(row=0, column=0)
    username = tk.StringVar()
    tk.Entry(tk_window, textvariable=username).grid(row=0, column=1)
    # password label and password entry box
    tk.Label(tk_window, text="Password").grid(row=1, column=0)
    password = tk.StringVar()
    tk.Entry(tk_window, textvariable=password).grid(row=1, column=1)
    # login button
    tk.Button(tk_window, text="Login", command=validate).grid(row=4, column=0)
    tk_window.mainloop()


def sheets_and_order_setup():  # ----------Main Process----------
    global browser
    browser = setup()
    gui_app("Enter Chillisoft details")
    pyautogui.confirm("Click OK to proceed")
    persons = get_profiles()
    wb = Workbook()
    # Sets up the Excel spreadsheet
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    sheet.title = "Nod32"
    wb.create_sheet("IntSec&Smart")
    wb.create_sheet("Multi")
    wb.create_sheet("CyberSecMac")
    wb.create_sheet("Other")
    # ----
    get_info(wb, persons)
    wb.save(file_location)
    print("--Completed, You May Now Close This Window--")
    raise SystemExit


def setup():  # creates the selenium browser, without it loading up
    option = webdriver.ChromeOptions()
    option.add_argument('headless')
    # , options=option
    browserValue = webdriver.Chrome('chromedriver.exe', options=option)
    return browserValue


def get_profiles():  # gets all the links to each of the users
    global browser
    browser.get(postVerifiedUrl)
    soup = bs4.BeautifulSoup(browser.page_source, 'lxml')
    persons = []
    for links in soup.select("#jax > div.uk-hidden-small > table > tbody > tr > td:nth-child(7) > a"):
        persons.append(links.get('href'))
    return persons


def get_info(wb, persons):  # get and add each user to the excel spreadsheet, based on their type of license
    global browser
    print("Getting info")
    nod32_count, int_sec_count, multi_count, mac_count, other_count = 0, 0, 0, 0, 0
    for index, person in enumerate(persons):
        browser.get(chillisoft_prefix + person)
        profile_html_parse = bs4.BeautifulSoup(browser.page_source, 'lxml')

        license_type = profile_html_parse.select("div > div.uk-width-7-10")[0].get_text()

        if license_type == "ESET NOD32 Antivirus":
            add_to_sheet(wb["Nod32"], index, profile_html_parse, nod32_count)
            nod32_count += 1
        elif license_type == "ESET Internet Security" \
                or license_type == "ESET Smart Security":
            add_to_sheet(wb["IntSec&Smart"], index, profile_html_parse, int_sec_count)
            int_sec_count += 1
        elif license_type == "ESET Multi-Device Security 5-Pack":
            add_to_sheet(wb["Multi"], index, profile_html_parse, multi_count)
            multi_count += 1
        elif license_type == "ESET Cyber Security for Mac":
            add_to_sheet(wb["CyberSecMac"], index, profile_html_parse, mac_count)
            mac_count += 1
        else:
            add_to_sheet(wb["Other"], index, profile_html_parse, other_count)
            other_count += 1


def add_to_sheet(sheet, index, soup, offset):  # after determining the type of licence, adds the values to the sheet --- note the qty and date values may be subject to change, based on updates to the website.
    sheet['A' + str(1 + offset)] = soup.select("#licenceeTab > div > div:nth-child(3)")[0].get_text()
    sheet['B' + str(1 + offset)] = soup.select("#licenceeTab > div > div:nth-child(5)")[0].get_text()
    sheet['C' + str(1 + offset)] = soup.select("div > div.uk-width-3-10.uk-text-right")[0].get_text()  # gets the qty
    sheet['D' + str(1 + offset)] = soup.select("div > div.uk-width-1-1")[5].get_text()  # gets the expiry date

    print("Added customer: " + str(index) + " | " + str(sheet['A' + str(1 + offset)].value) + " " +
          str(sheet['B' + str(1 + offset)].value) + " " + str(sheet['C' + str(1 + offset)].value) + " Expires:" + str(sheet['D' + str(1 + offset)].value))


def validate():  # checks the username and password works.
    global username
    global password
    global browser
    global tk_window
    tk_window.destroy()
    if len(bs4.BeautifulSoup(browser.page_source, 'lxml').select(
            "#jax > div.uk-hidden-small > table > thead > tr > th:nth-child(1)")) > 0:
        print("Sign in successful")
    else:
        login(username.get(), password.get())
        if len(bs4.BeautifulSoup(browser.page_source, 'lxml').select("body > section > div > div > p")) > 0:
            print("Unsuccessful sign in - try again")
            gui_app("Details unsuccessful")


def login(usr, pswd):  # proceeds to attempt to log the user in with the provided keys
    global browser
    login_id = "LoginName"
    password_id = "Password"
    submit_id = "sendbutton2"
    browser.get(login_url)
    if browser.find_element_by_id(login_id):
        item = browser.find_element_by_id(login_id)
        item.click()
        item.clear()
        item.send_keys(usr)

    if browser.find_element_by_id(password_id):
        item = browser.find_element_by_id(password_id)
        item.click()
        item.clear()
        item.send_keys(pswd)
        submit_btn = browser.find_element_by_name(submit_id)
        submit_btn.click()
    print("Signed in")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    run()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
